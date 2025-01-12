from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
import time

options = Options()
options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
driver = webdriver.Chrome( options=options)

employees = []
with open("people.csv", "r", encoding="utf-8") as file:
    next(file)
    for line in file:
        parts = line.strip().split(",")
        if len(parts) >= 3:
            full_name = (parts[2] + " " + parts[3]).strip().lower()
            employees.append(full_name)
        else:
            print(f"error: {line.strip()}")

print(employees)
url = "https://emn178.github.io/online-tools/crc32.html"
crc32_results = {}

driver.get(url)
time.sleep(2)

for employee in employees:
    try:
        input_field = driver.find_element(By.ID, "input")
        input_field.clear()
        input_field.send_keys(employee)
        
        WebDriverWait(driver, 10).until(
            EC.text_to_be_present_in_element_value((By.ID, "output"), "")
        )
        result_text = driver.find_element(By.ID, "output").get_attribute("value")

        # print(f"CRC32 for {employee}: {result_text}")
        crc32_results[employee] = result_text
    except Exception as e:
        print(f"error {employee}: {e}")
        crc32_results[employee] = "Unknown"

print(crc32_results)
driver.quit()

wb = load_workbook("salary.xlsx")
sheet = wb.active

output_wb = Workbook()
output_ws = output_wb.active
output_ws.append(["CRC32 Code", "Total Salary"])

for row in sheet.iter_rows(min_row=2, max_row=739):
    # print(row)
    name = row[0].value
    salary = row[1].value
    print(name, salary)
    encoded_name = crc32_results.get(name, "Unknown")
    output_ws.append([encoded_name, salary])

output_wb.save("updated_salary.xlsx")